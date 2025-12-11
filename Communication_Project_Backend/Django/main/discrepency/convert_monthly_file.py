import os
import pandas as pd
from openpyxl import load_workbook
import pdb

from main.settings import  BASE_DIR
# Define folder paths
input_folder = r"Z:\Santhosh D\Website_Conversion\Input_Folder"
# output_folder = r"Z:\Santhosh D\Website_Conversion\Output_Folder"


# june_file_path = r"Z:\Santhosh D\Website_Conversion\June_SR1_with_Remarks (1).xlsx"

# # Template for target columns
# june_data_template = pd.read_excel(june_file_path, sheet_name='Substation_Summary')
# target_columns = june_data_template.columns

target_columns = ['Substation', 'Substation_Name', 'Voltage_Level',
       'No_of_Non_Reporting_Points', 'No_of_Points',
       'Percentage_Number_of_Points_Non_Reporting (%)']

def transFormInputFile(monhtly_df , filename , current_month,prev_month):
    try:
        # Define file paths
        # input_file_path = os.path.join(input_folder, filename)
        directory = os.path.join(BASE_DIR , 'FILES' )

        if not os.path.exists(directory):
            # Create the directory if it doesn't exist
            os.makedirs(directory)

        output_file_path = os.path.join(directory , filename)
        formatted_data = {}
        for sheet_name, df in monhtly_df.items():
            formatted_df = df.reindex(columns=target_columns)
            formatted_data[sheet_name] = formatted_df
        
        # Save reformatted data to output file
        with pd.ExcelWriter(output_file_path) as writer:
            for sheet_name, formatted_df in formatted_data.items():
                formatted_df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        
        # Step 2: Modify the "Station_Summary" sheet
        workbook = load_workbook(output_file_path)
        if 'Station_Summary' in workbook.sheetnames:
            station_summary_df = pd.read_excel(output_file_path, sheet_name='Station_Summary')
            if 'Substation' in station_summary_df.columns:
                station_summary_df = station_summary_df.rename(columns={'Substation': 'S.No'})
                station_summary_df['S.No'] = range(1, len(station_summary_df) + 1)
            
            with pd.ExcelWriter(output_file_path, mode='a', if_sheet_exists='replace', engine='openpyxl') as writer:
                station_summary_df.to_excel(writer, sheet_name='Station_Summary', index=False)

        # Additional operations for all other sheets
        new_columns = [
            "ICCP_IOA", "SubStation", "Voltage Level", "ELEMENT_DESCRIPTION",
            "ELEMENT_CATEGORY", "Metric_Type", "No_of_Points",
            current_month+"_Non_Availability_Percentage", prev_month+"_Non_Availability_Percentage", "Latest Month Remarks", "Previous Month Remarks"
        ]

        with pd.ExcelWriter(output_file_path) as writer:
            for sheet_name, df in monhtly_df.items():
                if sheet_name != 'Station_Summary':
                    filtered_df = df[~df.apply(lambda row: row.astype(str).str.contains('Completely Not Reporting|Intermittent Data|ICCP_IOA').any(), axis=1)]
                    filtered_df.columns = new_columns[:len(filtered_df.columns)]
                    modified_df = filtered_df.copy()

                    modified_dynamic_cols = [col for col in modified_df.columns if "Non_Availability_Percentage" in col]
                    
                    grouped_df = modified_df.groupby("ICCP_IOA", as_index=False).agg({
                        **{col: "sum" for col in modified_dynamic_cols}
                    })
                else:
                    result_df = df.drop(df.columns[5:9], axis=1)
                    # Identify the dynamic column (e.g., the one with "Completely_Not_Reporting_Points")
                    dynamic_col = [col for col in result_df.columns if "Completely_Not_Reporting_Points" in col]

                    # Group by Substation_Name
                    grouped_df = result_df.groupby("Substation_Name", as_index=False).agg({
                        "Voltage_Level": "mean",
                        "No_of_Points": "sum",
                        **{col: "sum" for col in dynamic_col}
                    })

                grouped_df.reset_index(drop=True, inplace=True)
                grouped_df.to_excel(writer, sheet_name=sheet_name, index=False, header=True)

        final_df =  pd.read_excel(output_file_path, sheet_name=None) 
        # now remove the file from folder , no longer needed
        if os.path.exists(output_file_path):
            os.remove(output_file_path)
        return final_df

    except Exception as e:
        print(e)
        return None
    
# # Iterate over all files in the input folder
# for filename in os.listdir(input_folder):
#     if filename.endswith(".xlsx"):  # Process only Excel files
#         # Define file paths
#         input_file_path = os.path.join(input_folder, filename)
#         output_file_path = os.path.join(output_folder, filename)
        
#         # Step 1: Initial reformatting from October to June format
#         monhtly_df = pd.read_excel(input_file_path, None)  # Load all sheets
        
#         formatted_data = {}
#         for sheet_name, df in monhtly_df.items():
#             formatted_df = df.reindex(columns=target_columns)
#             formatted_data[sheet_name] = formatted_df
        
#         # Save reformatted data to output file
#         with pd.ExcelWriter(output_file_path) as writer:
#             for sheet_name, formatted_df in formatted_data.items():
#                 formatted_df.to_excel(writer, sheet_name=sheet_name, index=False)
        
#         print(f"Initial reformatting completed for {filename}.")

#         # Step 2: Modify the "Station_Summary" sheet
#         workbook = load_workbook(output_file_path)
#         if 'Station_Summary' in workbook.sheetnames:
#             station_summary_df = pd.read_excel(output_file_path, sheet_name='Station_Summary')
#             if 'Substation' in station_summary_df.columns:
#                 station_summary_df = station_summary_df.rename(columns={'Substation': 'S.No'})
#                 station_summary_df['S.No'] = range(1, len(station_summary_df) + 1)
            
#             with pd.ExcelWriter(output_file_path, mode='a', if_sheet_exists='replace', engine='openpyxl') as writer:
#                 station_summary_df.to_excel(writer, sheet_name='Station_Summary', index=False)

#             print(f"'Station_Summary' sheet updated for {filename}.")
#         else:
#             print(f"'Station_Summary' sheet not found in {filename}.")

#         # Additional operations for all other sheets
#         new_columns = [
#             "ICCP_IOA", "SubStation", "Voltage Level", "ELEMENT_DESCRIPTION",
#             "ELEMENT_CATEGORY", "Metric_Type", "No_of_Points",
#             "October_Non_Availability_Percentage", "September_Non_Availability_Percentage", "Latest Month Remarks", "Previous Month Remarks"
#         ]

#         with pd.ExcelWriter(output_file_path) as writer:
#             for sheet_name in pd.ExcelFile(input_file_path).sheet_names:
#                 df = pd.read_excel(input_file_path, sheet_name=sheet_name)
#                 if sheet_name != 'Station_Summary':
#                     filtered_df = df[~df.apply(lambda row: row.astype(str).str.contains('Completely Not Reporting|Intermittent Data|ICCP_IOA').any(), axis=1)]
#                     filtered_df.columns = new_columns[:len(filtered_df.columns)]
#                     modified_df = filtered_df.copy()
#                     grouped_df = modified_df.groupby("ICCP_IOA", as_index=False).agg({
#                         "October_Non_Availability_Percentage": "sum",
#                         "September_Non_Availability_Percentage": "sum",
#                     })
#                 else:
#                     grouped_df = df.drop(df.columns[5:9], axis=1)
                
#                 grouped_df.reset_index(drop=True, inplace=True)
#                 grouped_df.to_excel(writer, sheet_name=sheet_name, index=False, header=True)

#         print(f"Filtered and grouped data saved for {filename}.")
